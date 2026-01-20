"""
Tests for ExportService - multi-format export functionality
"""

import json
import os
import tempfile
import pytest

from src.shared.export_service import (
    ExportService,
    ExportFormat,
    parse_format_list,
    sanitize_csv_value,
    sanitize_store_for_csv,
    CSV_INJECTION_CHARS
)


# Sample store data for testing
SAMPLE_STORES = [
    {
        "store_id": "1001",
        "name": "Test Store 1",
        "address": "123 Main St",
        "city": "New York",
        "state": "NY",
        "zip_code": "10001",
        "latitude": 40.7128,
        "longitude": -74.0060,
        "phone": "212-555-0001"
    },
    {
        "store_id": "1002",
        "name": "Test Store 2",
        "address": "456 Oak Ave",
        "city": "Los Angeles",
        "state": "CA",
        "zip_code": "90001",
        "latitude": 34.0522,
        "longitude": -118.2437,
        "phone": "323-555-0002"
    }
]


class TestExportFormat:
    """Tests for ExportFormat enum"""

    def test_export_format_values(self):
        """Test that all export formats have correct values"""
        assert ExportFormat.JSON.value == "json"
        assert ExportFormat.CSV.value == "csv"
        assert ExportFormat.EXCEL.value == "excel"
        assert ExportFormat.GEOJSON.value == "geojson"

    def test_export_format_from_string(self):
        """Test creating ExportFormat from string"""
        assert ExportFormat("json") == ExportFormat.JSON
        assert ExportFormat("csv") == ExportFormat.CSV
        assert ExportFormat("excel") == ExportFormat.EXCEL
        assert ExportFormat("geojson") == ExportFormat.GEOJSON

    def test_export_format_from_string_method(self):
        """Test ExportFormat.from_string method"""
        assert ExportFormat.from_string("json") == ExportFormat.JSON
        assert ExportFormat.from_string("JSON") == ExportFormat.JSON
        assert ExportFormat.from_string("xlsx") == ExportFormat.EXCEL  # alias


class TestParseFormatList:
    """Tests for parse_format_list helper function"""

    def test_parse_single_format(self):
        """Test parsing a single format"""
        result = parse_format_list("json")
        assert result == [ExportFormat.JSON]

    def test_parse_multiple_formats(self):
        """Test parsing comma-separated formats"""
        result = parse_format_list("json,csv,excel")
        assert ExportFormat.JSON in result
        assert ExportFormat.CSV in result
        assert ExportFormat.EXCEL in result

    def test_parse_formats_with_spaces(self):
        """Test parsing formats with extra spaces"""
        result = parse_format_list("json, csv, excel")
        assert len(result) == 3

    def test_parse_invalid_format_ignored(self):
        """Test that invalid formats are ignored"""
        result = parse_format_list("json,invalid,csv")
        assert ExportFormat.JSON in result
        assert ExportFormat.CSV in result
        assert len(result) == 2

    def test_parse_empty_string(self):
        """Test parsing empty string returns empty list"""
        result = parse_format_list("")
        assert result == []

    def test_parse_all_invalid_returns_empty(self):
        """Test parsing all invalid formats returns empty list"""
        result = parse_format_list("invalid,unknown,bad")
        assert result == []


class TestExportServiceJSON:
    """Tests for JSON export functionality"""

    def test_export_json_to_file(self):
        """Test exporting JSON to file"""
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            output_path = f.name

        try:
            ExportService.export_stores(SAMPLE_STORES, ExportFormat.JSON, output_path)
            assert os.path.exists(output_path)

            with open(output_path, 'r') as f:
                data = json.load(f)
                assert len(data) == 2
                assert data[0]['store_id'] == '1001'
        finally:
            os.unlink(output_path)

    def test_export_json_preserves_all_fields(self):
        """Test that JSON export preserves all fields"""
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            output_path = f.name

        try:
            ExportService.export_stores(SAMPLE_STORES, ExportFormat.JSON, output_path)

            with open(output_path, 'r') as f:
                data = json.load(f)
                store = data[0]
                assert store['name'] == 'Test Store 1'
                assert store['latitude'] == 40.7128
                assert store['city'] == 'New York'
        finally:
            os.unlink(output_path)


class TestExportServiceCSV:
    """Tests for CSV export functionality"""

    def test_generate_csv_string_basic(self):
        """Test generating CSV string"""
        result = ExportService.generate_csv_string(SAMPLE_STORES)
        assert isinstance(result, str)

        lines = result.strip().split('\n')
        assert len(lines) == 3  # Header + 2 data rows

    def test_generate_csv_string_with_fieldnames(self):
        """Test generating CSV with specific fieldnames"""
        fieldnames = ['store_id', 'name', 'city']
        result = ExportService.generate_csv_string(SAMPLE_STORES, fieldnames)

        lines = result.strip().split('\n')
        header = lines[0]
        assert 'store_id' in header
        assert 'name' in header
        assert 'city' in header
        # Should not have other fields
        assert header.count(',') == 2

    def test_export_csv_to_file(self):
        """Test exporting CSV to file"""
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = f.name

        try:
            ExportService.export_stores(SAMPLE_STORES, ExportFormat.CSV, output_path)
            assert os.path.exists(output_path)

            with open(output_path, 'r') as f:
                content = f.read()
                assert 'Test Store 1' in content
                assert 'Test Store 2' in content
        finally:
            os.unlink(output_path)


class TestExportServiceGeoJSON:
    """Tests for GeoJSON export functionality"""

    def test_generate_geojson(self):
        """Test generating GeoJSON structure"""
        result = ExportService.generate_geojson(SAMPLE_STORES)

        assert result['type'] == 'FeatureCollection'
        assert 'features' in result
        assert len(result['features']) == 2

    def test_geojson_feature_structure(self):
        """Test that GeoJSON features have correct structure"""
        result = ExportService.generate_geojson(SAMPLE_STORES)
        feature = result['features'][0]

        assert feature['type'] == 'Feature'
        assert 'geometry' in feature
        assert 'properties' in feature

        # Check geometry
        assert feature['geometry']['type'] == 'Point'
        assert feature['geometry']['coordinates'] == [-74.0060, 40.7128]  # [lng, lat]

    def test_geojson_properties_include_all_fields(self):
        """Test that GeoJSON properties include all store fields"""
        result = ExportService.generate_geojson(SAMPLE_STORES)
        properties = result['features'][0]['properties']

        assert properties['store_id'] == '1001'
        assert properties['name'] == 'Test Store 1'
        assert properties['city'] == 'New York'

    def test_geojson_skips_stores_without_coordinates(self):
        """Test GeoJSON skips stores without valid coordinates"""
        stores_no_coords = [
            {'store_id': '1', 'name': 'No Coords Store'},
            {'store_id': '2', 'name': 'With Coords', 'latitude': 40.0, 'longitude': -74.0}
        ]
        result = ExportService.generate_geojson(stores_no_coords)

        # Only the store with coordinates should be included
        assert len(result['features']) == 1
        assert result['features'][0]['properties']['name'] == 'With Coords'

    def test_geojson_handles_zero_coordinates(self):
        """Test GeoJSON correctly handles stores at equator (lat=0) or prime meridian (lng=0)"""
        stores_with_zeros = [
            # Equator store (lat=0)
            {'store_id': '1', 'name': 'Equator Store', 'latitude': 0, 'longitude': 100.0},
            # Prime meridian store (lng=0)
            {'store_id': '2', 'name': 'Prime Meridian Store', 'latitude': 50.0, 'longitude': 0},
            # Origin point (lat=0, lng=0) - Null Island
            {'store_id': '3', 'name': 'Null Island', 'latitude': 0, 'longitude': 0},
            # Normal store for comparison
            {'store_id': '4', 'name': 'Normal Store', 'latitude': 40.0, 'longitude': -74.0}
        ]
        result = ExportService.generate_geojson(stores_with_zeros)

        # All 4 stores should be included
        assert len(result['features']) == 4
        
        # Verify equator store
        equator = result['features'][0]
        assert equator['properties']['name'] == 'Equator Store'
        assert equator['geometry']['coordinates'] == [100.0, 0]
        
        # Verify prime meridian store
        prime = result['features'][1]
        assert prime['properties']['name'] == 'Prime Meridian Store'
        assert prime['geometry']['coordinates'] == [0, 50.0]
        
        # Verify null island
        null_island = result['features'][2]
        assert null_island['properties']['name'] == 'Null Island'
        assert null_island['geometry']['coordinates'] == [0, 0]

    def test_geojson_fallback_coordinate_field_names(self):
        """Test GeoJSON correctly handles alternative coordinate field names"""
        stores_alt_fields = [
            # Using 'lat' and 'lng' instead of 'latitude' and 'longitude'
            {'store_id': '1', 'name': 'Store 1', 'lat': 40.0, 'lng': -74.0},
            # Using 'lat' and 'lon'
            {'store_id': '2', 'name': 'Store 2', 'lat': 50.0, 'lon': -120.0},
            # Using 'lat' with 0 value
            {'store_id': '3', 'name': 'Store 3', 'lat': 0, 'lng': 100.0},
            # Primary field names take precedence even if 0
            {'store_id': '4', 'name': 'Store 4', 'latitude': 0, 'lat': 50.0, 'longitude': 0, 'lng': 100.0}
        ]
        result = ExportService.generate_geojson(stores_alt_fields)

        # All 4 stores should be included
        assert len(result['features']) == 4
        
        # Store with lat/lng
        assert result['features'][0]['geometry']['coordinates'] == [-74.0, 40.0]
        
        # Store with lat/lon
        assert result['features'][1]['geometry']['coordinates'] == [-120.0, 50.0]
        
        # Store with zero lat
        assert result['features'][2]['geometry']['coordinates'] == [100.0, 0]
        
        # Primary fields take precedence (latitude/longitude should be used, not lat/lng)
        assert result['features'][3]['geometry']['coordinates'] == [0, 0]

    def test_export_geojson_to_file(self):
        """Test exporting GeoJSON to file"""
        with tempfile.NamedTemporaryFile(suffix='.geojson', delete=False) as f:
            output_path = f.name

        try:
            ExportService.export_stores(SAMPLE_STORES, ExportFormat.GEOJSON, output_path)
            assert os.path.exists(output_path)

            with open(output_path, 'r') as f:
                data = json.load(f)
                assert data['type'] == 'FeatureCollection'
        finally:
            os.unlink(output_path)


class TestExportServiceExcel:
    """Tests for Excel export functionality"""

    def test_generate_excel_bytes(self):
        """Test generating Excel bytes"""
        result = ExportService.generate_excel_bytes(SAMPLE_STORES, 'TestSheet')
        assert isinstance(result, bytes)
        # Check for Excel file signature (PK for zip format)
        assert result[:2] == b'PK'

    def test_generate_excel_bytes_with_fieldnames(self):
        """Test generating Excel with specific fieldnames"""
        fieldnames = ['store_id', 'name', 'city']
        result = ExportService.generate_excel_bytes(SAMPLE_STORES, 'Test', fieldnames)
        assert isinstance(result, bytes)

    def test_generate_multi_sheet_excel(self):
        """Test generating multi-sheet Excel workbook"""
        retailer_data = {
            'retailer1': SAMPLE_STORES,
            'retailer2': SAMPLE_STORES[:1]
        }

        result = ExportService.generate_multi_sheet_excel(retailer_data)
        assert isinstance(result, bytes)
        assert result[:2] == b'PK'

    def test_export_excel_to_file(self):
        """Test exporting Excel to file"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        try:
            ExportService.export_stores(SAMPLE_STORES, ExportFormat.EXCEL, output_path)
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0
        finally:
            os.unlink(output_path)

    def test_generate_multi_sheet_excel_empty_retailer_skipped(self):
        """Test that empty retailers are skipped in multi-sheet Excel"""
        retailer_data = {
            'retailer1': SAMPLE_STORES,
            'empty_retailer': []
        }

        result = ExportService.generate_multi_sheet_excel(retailer_data)
        assert isinstance(result, bytes)

    def test_generate_multi_sheet_excel_all_empty_retailers(self):
        """Test that multi-sheet Excel with all empty retailers doesn't crash"""
        retailer_data = {
            'retailer1': [],
            'retailer2': [],
            'retailer3': []
        }

        # Should not raise ValueError about no worksheets
        result = ExportService.generate_multi_sheet_excel(retailer_data)
        assert isinstance(result, bytes)


class TestExportServiceIntegration:
    """Integration tests for ExportService"""

    def test_export_stores_creates_directory(self):
        """Test that export_stores creates output directory if needed"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'subdir', 'output.json')
            ExportService.export_stores(SAMPLE_STORES, ExportFormat.JSON, output_path)

            assert os.path.exists(output_path)

    def test_export_stores_with_config(self):
        """Test export_stores with retailer config containing output_fields"""
        config = {
            'output_fields': ['store_id', 'name', 'city']
        }

        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = f.name

        try:
            ExportService.export_stores(SAMPLE_STORES, ExportFormat.CSV, output_path, config)
            with open(output_path, 'r') as f:
                content = f.read()
                # Should have limited fields
                lines = content.strip().split('\n')
                assert lines[0].count(',') == 2
        finally:
            os.unlink(output_path)

    def test_all_formats_export_successfully(self):
        """Test that all formats can export successfully"""
        formats = [ExportFormat.JSON, ExportFormat.CSV, ExportFormat.EXCEL, ExportFormat.GEOJSON]
        extensions = ['json', 'csv', 'xlsx', 'geojson']

        with tempfile.TemporaryDirectory() as tmpdir:
            for fmt, ext in zip(formats, extensions):
                output_path = os.path.join(tmpdir, f'test.{ext}')
                ExportService.export_stores(SAMPLE_STORES, fmt, output_path)
                assert os.path.exists(output_path), f"Failed to create {fmt.value} file"
                assert os.path.getsize(output_path) > 0, f"{fmt.value} file is empty"

    def test_export_empty_stores_logs_warning(self):
        """Test that exporting empty stores doesn't fail"""
        with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as f:
            output_path = f.name

        try:
            # Should not raise, just log warning
            ExportService.export_stores([], ExportFormat.JSON, output_path)
            # File may or may not exist (implementation choice)
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


class TestFormulaInjectionProtection:
    """Tests for CSV/Excel formula injection protection (#73)"""

    def test_csv_injection_chars_constant(self):
        """Test that CSV_INJECTION_CHARS includes all dangerous characters"""
        assert '=' in CSV_INJECTION_CHARS
        assert '+' in CSV_INJECTION_CHARS
        assert '-' in CSV_INJECTION_CHARS
        assert '@' in CSV_INJECTION_CHARS
        assert '\t' in CSV_INJECTION_CHARS
        assert '\r' in CSV_INJECTION_CHARS
        assert '\n' in CSV_INJECTION_CHARS

    def test_sanitize_csv_value_equals_sign(self):
        """Test that values starting with = are sanitized"""
        assert sanitize_csv_value('=1+1') == "'=1+1"
        assert sanitize_csv_value('=SUM(A1:A10)') == "'=SUM(A1:A10)"
        assert sanitize_csv_value('=cmd|/c calc') == "'=cmd|/c calc"

    def test_sanitize_csv_value_plus_sign(self):
        """Test that values starting with + are sanitized"""
        assert sanitize_csv_value('+1+1') == "'+1+1"
        assert sanitize_csv_value('+cmd|/c calc') == "'+cmd|/c calc"

    def test_sanitize_csv_value_minus_sign(self):
        """Test that values starting with - are sanitized"""
        assert sanitize_csv_value('-1+1') == "'-1+1"
        assert sanitize_csv_value('-cmd|/c calc') == "'-cmd|/c calc"

    def test_sanitize_csv_value_at_sign(self):
        """Test that values starting with @ are sanitized"""
        assert sanitize_csv_value('@SUM(A1:A10)') == "'@SUM(A1:A10)"

    def test_sanitize_csv_value_tab_carriage_newline(self):
        """Test that values starting with tab, CR, or newline are sanitized"""
        assert sanitize_csv_value('\ttest') == "'\ttest"
        assert sanitize_csv_value('\rtest') == "'\rtest"
        assert sanitize_csv_value('\ntest') == "'\ntest"

    def test_sanitize_csv_value_safe_strings(self):
        """Test that safe strings are not modified"""
        assert sanitize_csv_value('normal text') == 'normal text'
        assert sanitize_csv_value('123 Main St') == '123 Main St'
        assert sanitize_csv_value('test@example.com') == 'test@example.com'
        assert sanitize_csv_value('(123) 456-7890') == '(123) 456-7890'

    def test_sanitize_csv_value_none(self):
        """Test that None values are unchanged"""
        assert sanitize_csv_value(None) is None

    def test_sanitize_csv_value_non_string(self):
        """Test that non-string values are unchanged"""
        assert sanitize_csv_value(123) == 123
        assert sanitize_csv_value(45.67) == 45.67
        assert sanitize_csv_value(True) is True
        assert sanitize_csv_value([1, 2, 3]) == [1, 2, 3]

    def test_sanitize_csv_value_empty_string(self):
        """Test that empty strings are unchanged"""
        assert sanitize_csv_value('') == ''

    def test_sanitize_store_for_csv(self):
        """Test that all string values in a store dict are sanitized"""
        dangerous_store = {
            'name': '=MALICIOUS()',
            'address': '+evil command',
            'city': 'Safe City',
            'zip': '12345',
            'phone': '@SUM(A1:A10)'
        }
        sanitized = sanitize_store_for_csv(dangerous_store)
        
        assert sanitized['name'] == "'=MALICIOUS()"
        assert sanitized['address'] == "'+evil command"
        assert sanitized['city'] == 'Safe City'  # unchanged
        assert sanitized['zip'] == '12345'  # unchanged
        assert sanitized['phone'] == "'@SUM(A1:A10)"

    def test_csv_export_sanitizes_dangerous_values(self):
        """Test that CSV export properly sanitizes formula injection attempts"""
        dangerous_stores = [
            {
                'store_id': '1001',
                'name': '=1+1',
                'address': '+dangerous',
                'phone': '@SUM(A1:A10)'
            }
        ]
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
            output_path = f.name
        
        try:
            ExportService.export_stores(
                dangerous_stores,
                ExportFormat.CSV,
                output_path
            )
            
            # Read the CSV and verify sanitization
            with open(output_path, 'r', encoding='utf-8') as f:
                content = f.read()
                assert "'=1+1" in content
                assert "'+dangerous" in content
                assert "'@SUM" in content
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_excel_export_sanitizes_dangerous_values(self):
        """Test that Excel export properly sanitizes formula injection attempts"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        
        dangerous_stores = [
            {
                'store_id': '1001',
                'name': '=1+1',
                'address': '+dangerous',
                'phone': '@SUM(A1:A10)'
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            ExportService.export_stores(
                dangerous_stores,
                ExportFormat.EXCEL,
                output_path
            )
            
            # Load the Excel file and verify sanitization
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Row 1 is headers, row 2 is data
            assert ws['B2'].value == "'=1+1"  # name column
            assert ws['C2'].value == "'+dangerous"  # address column
            assert ws['D2'].value == "'@SUM(A1:A10)"  # phone column
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_generate_csv_string_sanitizes_values(self):
        """Test that generate_csv_string sanitizes dangerous values"""
        dangerous_stores = [
            {
                'name': '=MALICIOUS()',
                'value': '+123'
            }
        ]
        
        csv_string = ExportService.generate_csv_string(
            dangerous_stores,
            fieldnames=['name', 'value']
        )
        
        assert "'=MALICIOUS()" in csv_string
        assert "'+123" in csv_string

    def test_generate_excel_bytes_sanitizes_values(self):
        """Test that generate_excel_bytes sanitizes dangerous values"""
        try:
            from openpyxl import load_workbook
            from io import BytesIO
        except ImportError:
            pytest.skip("openpyxl not available")
        
        dangerous_stores = [
            {
                'name': '=MALICIOUS()',
                'value': '-evil'
            }
        ]
        
        excel_bytes = ExportService.generate_excel_bytes(
            dangerous_stores,
            fieldnames=['name', 'value']
        )
        
        # Load from bytes and verify sanitization
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active
        
        assert ws['A2'].value == "'=MALICIOUS()"
        assert ws['B2'].value == "'-evil"
