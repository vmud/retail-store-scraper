"""
Export Service - Multi-format export functionality for store data.

Supports exporting to JSON, CSV, Excel (.xlsx), and GeoJSON formats.
"""

import csv
import json
import logging
from enum import Enum
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from src.shared.constants import EXPORT

try:
    from openpyxl import Workbook  # pylint: disable=import-error
    from openpyxl.styles import Font, Alignment  # pylint: disable=import-error
    from openpyxl.utils import get_column_letter  # pylint: disable=import-error
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportFormat(Enum):
    """Supported export formats."""
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    GEOJSON = "geojson"

    @classmethod
    def from_string(cls, value: str) -> "ExportFormat":
        """Parse format from string, case-insensitive."""
        value_lower = value.lower().strip()
        # Handle xlsx as alias for excel
        if value_lower == "xlsx":
            value_lower = "excel"
        for fmt in cls:
            if fmt.value == value_lower:
                return fmt
        raise ValueError(f"Unknown export format: {value}")


# Characters that can trigger formula injection in spreadsheet applications
CSV_INJECTION_CHARS = ('=', '+', '-', '@', '\t', '\r', '\n')


def sanitize_csv_value(value: Any) -> Any:
    """Sanitize a value for CSV export to prevent formula injection (#73).

    Spreadsheet applications like Excel interpret cells starting with
    =, +, -, @, tab, or carriage return as formulas. This can be exploited
    for CSV injection attacks.

    The mitigation is to prefix potentially dangerous values with a single quote,
    which Excel treats as a text prefix.

    Args:
        value: The value to sanitize

    Returns:
        The sanitized value (string prefixed with quote if dangerous, else unchanged)
    """
    if value is None:
        return value
    if not isinstance(value, str):
        return value
    if value and value[0] in CSV_INJECTION_CHARS:
        # Exception: Don't sanitize negative numbers (e.g., -92.0963940)
        # These are safe and common in coordinate data
        if value[0] == '-':
            try:
                float(value)
                return value  # It's a valid negative number, don't sanitize
            except ValueError:
                pass  # Not a number, continue with sanitization
        return f"'{value}"
    return value


def sanitize_store_for_csv(store: Dict[str, Any]) -> Dict[str, Any]:
    """Sanitize all string values in a store dict for CSV export.

    Args:
        store: Store dictionary

    Returns:
        New dictionary with sanitized values
    """
    return {key: sanitize_csv_value(value) for key, value in store.items()}


class ExportService:
    """Service for exporting store data to various formats."""

    # Default fields if none specified
    DEFAULT_FIELDS = [
        'name', 'street_address', 'city', 'state', 'zip',
        'country', 'latitude', 'longitude', 'phone', 'url', 'scraped_at'
    ]

    @staticmethod
    def export_stores(
        stores: List[Dict[str, Any]],
        export_format: ExportFormat,
        output_path: str,
        retailer_config: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Export stores to a file in the specified format.

        Args:
            stores: List of store dictionaries
            export_format: Target format (JSON, CSV, EXCEL, GEOJSON)
            output_path: Path to save the output file
            retailer_config: Optional retailer config with output_fields
        """
        if not stores:
            logging.warning("No stores to export")
            return

        # Validate output path to prevent path traversal attacks
        path = Path(output_path)
        # Check for path traversal attempts
        if ".." in str(path) or ".." in str(path.resolve()):
            raise ValueError(f"Invalid output path: {output_path}. Path traversal not allowed.")

        path.parent.mkdir(parents=True, exist_ok=True)

        fieldnames = ExportService._get_fieldnames(stores, retailer_config)

        if export_format == ExportFormat.JSON:
            ExportService._save_json(stores, path)
        elif export_format == ExportFormat.CSV:
            ExportService._save_csv(stores, path, fieldnames)
        elif export_format == ExportFormat.EXCEL:
            ExportService._save_excel(stores, path, fieldnames)
        elif export_format == ExportFormat.GEOJSON:
            ExportService._save_geojson(stores, path)

        logging.info(f"Exported {len(stores)} stores to {export_format.value.upper()}: {output_path}")

    @staticmethod
    def _get_fieldnames(
        stores: List[Dict[str, Any]],
        retailer_config: Optional[Dict[str, Any]] = None,
        sample_size: int = EXPORT.FIELD_SAMPLE_SIZE
    ) -> List[str]:
        """
        Get field names from config or infer from data.

        Takes union of fieldnames across first N records for completeness (#146).

        Args:
            stores: List of store dictionaries
            retailer_config: Optional retailer config with output_fields
            sample_size: Number of stores to sample for field discovery (default: 100)

        Returns:
            List of field names to include in export (sorted for determinism)
        """
        # Try to get from config
        if retailer_config and 'output_fields' in retailer_config:
            return retailer_config['output_fields']

        # Infer from first N stores (union of all fields) (#146)
        if stores:
            all_fields = set()
            for store in stores[:sample_size]:
                all_fields.update(store.keys())
            # Return sorted for deterministic column order
            return sorted(all_fields)

        # Fallback to defaults
        return ExportService.DEFAULT_FIELDS

    @staticmethod
    def _save_json(stores: List[Dict[str, Any]], path: Path) -> None:
        """Save stores to JSON file."""
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(stores, f, indent=2, ensure_ascii=False)

    @staticmethod
    def _save_csv(
        stores: List[Dict[str, Any]],
        path: Path,
        fieldnames: List[str]
    ) -> None:
        """Save stores to CSV file with formula injection protection (#73)."""
        # Sanitize stores to prevent CSV injection
        sanitized_stores = [sanitize_store_for_csv(store) for store in stores]
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(sanitized_stores)

    @staticmethod
    def _save_excel(
        stores: List[Dict[str, Any]],
        path: Path,
        fieldnames: List[str],
        sheet_name: str = "Stores"
    ) -> None:
        """Save stores to Excel file with formatting and formula injection protection (#5 review feedback)."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        # Sanitize stores to prevent formula injection (same as CSV)
        sanitized_stores = [sanitize_store_for_csv(store) for store in stores]

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write header row with bold formatting
        header_font = Font(bold=True)
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Write data rows (using sanitized data)
        for row_idx, store in enumerate(sanitized_stores, start=2):
            for col_idx, field in enumerate(fieldnames, start=1):
                value = store.get(field, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths
        for col_idx, field in enumerate(fieldnames, start=1):
            max_length = len(str(field))
            for row_idx in range(2, min(len(stores) + 2, 100)):  # Sample first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            # Cap at reasonable width
            adjusted_width = min(max_length + 2, EXPORT.EXCEL_MAX_COLUMN_WIDTH)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        wb.save(path)

    @staticmethod
    def _save_geojson(stores: List[Dict[str, Any]], path: Path) -> None:
        """Save stores to GeoJSON file."""
        geojson = ExportService.generate_geojson(stores)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(geojson, f, indent=2, ensure_ascii=False)

    @staticmethod
    def generate_geojson(stores: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate GeoJSON FeatureCollection from stores.

        Args:
            stores: List of store dictionaries

        Returns:
            GeoJSON FeatureCollection dictionary
        """
        features = []
        skipped = 0

        for store in stores:
            # Get coordinates - try multiple field names
            # Use explicit None checks to handle 0 values (equator/prime meridian)
            lat = store.get('latitude') if store.get('latitude') is not None else store.get('lat')
            lng = store.get('longitude') if store.get('longitude') is not None else (
                store.get('lng') if store.get('lng') is not None else store.get('lon')
            )

            # Skip stores without valid coordinates
            if lat is None or lng is None:
                skipped += 1
                continue

            try:
                lat_float = float(lat)
                lng_float = float(lng)
            except (ValueError, TypeError):
                skipped += 1
                continue

            # Skip invalid coordinates
            if not (-90 <= lat_float <= 90) or not (-180 <= lng_float <= 180):
                skipped += 1
                continue

            # Build feature with all store properties
            feature = {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    # GeoJSON uses [longitude, latitude] order
                    "coordinates": [lng_float, lat_float]
                },
                "properties": dict(store.items())
            }
            features.append(feature)

        if skipped > 0:
            logging.warning(f"Skipped {skipped} stores with missing or invalid coordinates")

        return {
            "type": "FeatureCollection",
            "features": features
        }

    @staticmethod
    def generate_excel_bytes(
        stores: List[Dict[str, Any]],
        sheet_name: str = "Stores",
        fieldnames: Optional[List[str]] = None
    ) -> bytes:
        """
        Generate Excel workbook in memory and return as bytes.

        Includes formula injection protection (#5 review feedback).

        Args:
            stores: List of store dictionaries
            sheet_name: Name for the worksheet
            fieldnames: Optional list of fields to include

        Returns:
            Excel file content as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        if fieldnames is None and stores:
            fieldnames = list(stores[0].keys())
        elif fieldnames is None:
            fieldnames = ExportService.DEFAULT_FIELDS

        # Sanitize stores to prevent formula injection (same as CSV)
        sanitized_stores = [sanitize_store_for_csv(store) for store in stores]

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write header row
        header_font = Font(bold=True)
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Write data rows (using sanitized data)
        for row_idx, store in enumerate(sanitized_stores, start=2):
            for col_idx, field in enumerate(fieldnames, start=1):
                value = store.get(field, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths
        for col_idx, field in enumerate(fieldnames, start=1):
            max_length = len(str(field))
            for row_idx in range(2, min(len(stores) + 2, 100)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, EXPORT.EXCEL_MAX_COLUMN_WIDTH)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_multi_sheet_excel(
        retailer_data: Dict[str, List[Dict[str, Any]]],
        config: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Generate multi-sheet Excel workbook with one sheet per retailer.

        Includes formula injection protection (#5 review feedback).

        Args:
            retailer_data: Dictionary mapping retailer names to store lists
            config: Optional config dict with per-retailer output_fields

        Returns:
            Excel file content as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        sheets_created = 0
        for retailer, stores in retailer_data.items():
            if not stores:
                continue

            # Get retailer-specific fieldnames
            retailer_config = config.get(retailer, {}) if config else {}
            fieldnames = retailer_config.get('output_fields')
            if not fieldnames:
                fieldnames = list(stores[0].keys())

            # Sanitize stores to prevent formula injection (same as CSV)
            sanitized_stores = [sanitize_store_for_csv(store) for store in stores]

            # Create sheet with capitalized retailer name
            sheet_name = retailer.title()[:EXPORT.EXCEL_SHEET_NAME_MAX]  # Excel sheet names max 31 chars
            ws = wb.create_sheet(title=sheet_name)
            sheets_created += 1

            # Write header row
            header_font = Font(bold=True)
            for col_idx, field in enumerate(fieldnames, start=1):
                cell = ws.cell(row=1, column=col_idx, value=field)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Write data rows (using sanitized data)
            for row_idx, store in enumerate(sanitized_stores, start=2):
                for col_idx, field in enumerate(fieldnames, start=1):
                    value = store.get(field, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-fit column widths
            for col_idx, field in enumerate(fieldnames, start=1):
                max_length = len(str(field))
                for row_idx in range(2, min(len(stores) + 2, 100)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, EXPORT.EXCEL_MAX_COLUMN_WIDTH)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = 'A2'

        # If no sheets were created (all retailers had empty data), create a placeholder sheet
        if sheets_created == 0:
            ws = wb.create_sheet(title="No Data")
            ws['A1'] = "No store data available for any retailer"
            ws['A1'].font = Font(bold=True)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_csv_string(
        stores: List[Dict[str, Any]],
        fieldnames: Optional[List[str]] = None
    ) -> str:
        """
        Generate CSV content as a string with formula injection protection (#73).

        Args:
            stores: List of store dictionaries
            fieldnames: Optional list of fields to include

        Returns:
            CSV content as string
        """
        if fieldnames is None and stores:
            fieldnames = list(stores[0].keys())
        elif fieldnames is None:
            fieldnames = ExportService.DEFAULT_FIELDS

        # Sanitize stores to prevent CSV injection
        sanitized_stores = [sanitize_store_for_csv(store) for store in stores]

        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(sanitized_stores)
        return output.getvalue()


def parse_format_list(format_string: str) -> List[ExportFormat]:
    """
    Parse comma-separated format string into list of ExportFormat.

    Args:
        format_string: Comma-separated format names (e.g., "json,csv,excel")

    Returns:
        List of ExportFormat enums
    """
    formats = []
    for fmt_str in format_string.split(','):
        fmt_str = fmt_str.strip()
        if fmt_str:
            try:
                formats.append(ExportFormat.from_string(fmt_str))
            except ValueError as e:
                logging.warning(str(e))
    return formats
